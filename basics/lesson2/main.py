import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
prod_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}


for product_row in range(2,prod_list.max_row + 1):
    supplier_name = prod_list.cell(product_row,4).value
    inventory = prod_list.cell(product_row,2).value
    price = prod_list.cell(product_row,3).value
    
    #calculation number of products per supplier
    if supplier_name in products_per_supplier:
       current_num_products = products_per_supplier.get(supplier_name)
       products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier")
        products_per_supplier[supplier_name] = 1


#calculation total value of inventory per supplier

if supplier_name in total_value_per_supplier:
    current_total_value = total_value_per_supplier.get(supplier_name)
    total_value_per_supplier[supplier_name] = current_total_value + inventory * price
else:   
    total_value_per_supplier[supplier_name] = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)